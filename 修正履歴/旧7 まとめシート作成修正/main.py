from __future__ import annotations

import sys
import traceback
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Any

import openpyxl  # テンプレ不要の新規作成

from io_fs import list_subfolders_sorted, circuit_csv_path, circuit_csv_path_root
from io_table import (
    ensure_sheet_by_index_and_rename,
    read_csv_cell_1based,
    read_csv_column_1based,
    col_letter,
)

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception:
    tk = None


JW_HEADERS = [
    "U相電源電圧(6)",
    "Vabs",
    "Vの線間電圧(実効値)",
    "Iabs",
    "Iの相電流(実効値)",
    "", "", "",
    "Vabs(3+6)",
    "Vの線間電圧(実効値)",
]

TRANSIENT_HEADERS = [
    "U相電源電圧(6)",
    "Iの実効値",
    "",
    "Vの実効値",
    "Vの線間電圧",
]


def _write_row(ws, row: int, values: List[str], start_col: int = 1):
    """1行分を左から順に書く（既存でも上書き）"""
    for j, v in enumerate(values):
        ws.cell(row=row, column=start_col + j, value=v)


def run(
    A: Path,
    output_xlsx: Optional[Path] = None,
    header_row: bool = True,
    read_to_write_map: Optional[List[Dict[str, Any]]] = None,
) -> Path:
    """
    ✅テンプレ不要版：
    - Excelは新規作成（openpyxl.Workbook）
    - jwシート/過渡シートのヘッダーを自動生成
    - 最後に「まとめ」シートを生成し、jw/過渡を転記
    """
    A = A.resolve()

    if read_to_write_map is None:
        read_to_write_map = [
            {"src_row": 8, "src_col": 4, "dst_col": 4, "label": "D8"},  # jw: D8 -> D
            {"src_row": 8, "src_col": 8, "dst_col": 2, "label": "H8"},  # jw: H8 -> B
        ]

    if output_xlsx is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_xlsx = Path.cwd() / f"無負荷特性_全自動化_out_{stamp}.xlsx"
    else:
        output_xlsx = output_xlsx.resolve()

    if not A.is_dir():
        raise FileNotFoundError(f"A folder not found: {A}")

    wb = openpyxl.Workbook()

    # Logging
    log_path = output_xlsx.with_suffix(".log.txt")
    log: List[str] = []
    log.append(f"Run at: {datetime.now().isoformat(timespec='seconds')}")
    log.append(f"A: {A}")
    log.append(f"Output: {output_xlsx}")
    log.append("")

    n_folders = list_subfolders_sorted(A)
    log.append(f"Found n folders: {len(n_folders)}")
    log.append("")

    # ★ヘッダーありなら n=1 は2行目から
    base_row_for_single = 2 if header_row else 1

    # 列貼り（過渡列ベタ貼り）の開始行
    start_row_for_columns = 2 if header_row else 1

    # 過渡シートのヘッダー（1行目/2行目）開始列：M列
    transient_header_start_col = 13  # M = 13

    # RMS式の参照範囲（固定指定）
    rms_r1 = 1283
    rms_r2 = 1538
    rms_N = 256

    # =========================================================
    # ★ 事前スキャン：mmax_global と mフォルダ名を確定
    #   （まとめシート・過渡シート番号を安定させるため）
    # =========================================================
    mmax_global = 0
    m_names_by_index: Dict[int, str] = {}  # m_idx(1-based) -> name

    for n_folder in n_folders:
        m_folders_tmp = list_subfolders_sorted(n_folder)
        mmax_global = max(mmax_global, len(m_folders_tmp))
        for i, mf in enumerate(m_folders_tmp, start=1):
            # 最初に見た名前を採用（通常、全nで同じ想定）
            if i not in m_names_by_index:
                m_names_by_index[i] = mf.name

    log.append(f"Detected mmax_global: {mmax_global}")
    log.append("")

    # =========================================================
    # 本処理
    # =========================================================
    for n_idx, n_folder in enumerate(n_folders, start=1):
        m_folders = list_subfolders_sorted(n_folder)
        log.append(f"[n={n_idx}/{len(n_folders)}] {n_folder.name} -> m folders: {len(m_folders)}")

        target_row_single = (base_row_for_single - 1) + n_idx  # headerあり: 2,3,4...

        for m_idx, m_folder in enumerate(m_folders, start=1):
            # -------------------------
            # sheet#1..mmax_global は「定常」(ABH / Original / ...)
            # -------------------------
            sheet_index = m_idx
            ws = ensure_sheet_by_index_and_rename(wb, sheet_index, desired_title=m_folder.name)

            # jwヘッダー
            if header_row:
                _write_row(ws, row=1, values=JW_HEADERS, start_col=1)

            # jwシートのA列に nフォルダ名
            ws.cell(row=target_row_single, column=1, value=n_folder.name)

            # -------------------------
            # jwシート：数式を自動入力
            #   C列: =B{r}/SQRT(2)*SQRT(3)
            #   E列: =D{r}/SQRT(2)
            #   I列: =B{r}+G{r}
            #   J列: =I{r}/SQRT(2)*SQRT(3)
            # -------------------------
            r = target_row_single
            ws.cell(row=r, column=3, value=f"=B{r}/SQRT(2)*SQRT(3)")   # C
            ws.cell(row=r, column=5, value=f"=D{r}/SQRT(2)")          # E
            ws.cell(row=r, column=9, value=f"=B{r}+G{r}")             # I
            ws.cell(row=r, column=10, value=f"=I{r}/SQRT(2)*SQRT(3)") # J

            # -------------------------
            # (A) jw/circuit.csv：単点貼り
            # -------------------------
            jw_csv = circuit_csv_path(n_folder, m_folder, jw_folder_name="jw", csv_name="circuit.csv")
            try:
                if not jw_csv.exists():
                    log.append(f"  [m={m_idx}] {m_folder.name}: MISSING {jw_csv}")
                else:
                    for item in read_to_write_map:
                        src_r = int(item["src_row"])
                        src_c = int(item["src_col"])
                        dst_c = int(item["dst_col"])
                        label = str(item.get("label", f"R{src_r}C{src_c}"))

                        val, raw = read_csv_cell_1based(jw_csv, row_1based=src_r, col_1based=src_c)
                        dstL = col_letter(dst_c)

                        ws.cell(row=target_row_single, column=dst_c, value=val if val is not None else None)

                        if val is None:
                            log.append(
                                f"  [m={m_idx}] {m_folder.name}: {label}(raw='{raw}') -> "
                                f"sheet#{sheet_index}('{ws.title}') {dstL}{target_row_single}=BLANK"
                            )
                        else:
                            log.append(
                                f"  [m={m_idx}] {m_folder.name}: {label}={val} -> "
                                f"sheet#{sheet_index}('{ws.title}') {dstL}{target_row_single}"
                            )

            except Exception as e:
                log.append(f"  [m={m_idx}] {m_folder.name}: ERROR (jw) {e}")
                log.append("    " + traceback.format_exc().replace("\n", "\n    "))

            # -------------------------
            # (B) m直下 circuit.csv：列貼り + 過渡シート
            # -------------------------
            root_csv = circuit_csv_path_root(m_folder, csv_name="circuit.csv")
            try:
                if not root_csv.exists():
                    log.append(f"  [m={m_idx}] {m_folder.name}: MISSING {root_csv}")
                else:
                    # ★過渡シートは「mmax_global+m_idx」で固定（全nで一貫）
                    transient_sheet_index = mmax_global + m_idx
                    transient_title = f"{m_folder.name}_過渡(V3+V6)"
                    ws_tr = ensure_sheet_by_index_and_rename(
                        wb, transient_sheet_index, desired_title=transient_title
                    )

                    # 過渡ヘッダー
                    if header_row:
                        _write_row(ws_tr, row=1, values=TRANSIENT_HEADERS, start_col=1)

                    base_dst = transient_header_start_col + 3 * (n_idx - 1)  # M + 3*(n-1)
                    dst_cols = [base_dst, base_dst + 1, base_dst + 2]

                    src_cols = [8, 17, 20]  # H, Q, T（1始まり）

                    summary_row = (2 if header_row else 1) + (n_idx - 1)

                    # 過渡：E列 = D*SQRT(3)
                    r2 = summary_row
                    ws_tr.cell(row=r2, column=5, value=f"=D{r2}*SQRT(3)")  # E

                    # A列に nフォルダ名（過渡）
                    ws_tr.cell(row=summary_row, column=1, value=n_folder.name)

                    # 1行目のM列から nフォルダ名を3列間隔
                    ws_tr.cell(row=1, column=base_dst, value=n_folder.name)

                    # 2行目のM列から「I」「V3」「V6」
                    ws_tr.cell(row=2, column=base_dst, value="I")
                    ws_tr.cell(row=2, column=base_dst + 1, value="V3")
                    ws_tr.cell(row=2, column=base_dst + 2, value="V6")

                    # B列 RMS（X=M/P/S...）
                    X = col_letter(base_dst)
                    ws_tr.cell(
                        row=summary_row,
                        column=2,
                        value=f"=SQRT(SUMSQ(${X}${rms_r1}:${X}${rms_r2})/{rms_N})",
                    )

                    # D列 RMS（Y=N/Q/T..., Z=O/R/U...）
                    Y = col_letter(base_dst + 1)
                    Z = col_letter(base_dst + 2)
                    ws_tr.cell(
                        row=summary_row,
                        column=4,
                        value=(
                            f"=SQRT(SUMSQ(${Y}${rms_r1}:${Y}${rms_r2})/{rms_N})"
                            f"+SQRT(SUMSQ(${Z}${rms_r1}:${Z}${rms_r2})/{rms_N})"
                        ),
                    )

                    # 列を丸ごと貼り
                    for src_c, dst_c in zip(src_cols, dst_cols):
                        values = read_csv_column_1based(root_csv, col_1based=src_c)
                        dstL = col_letter(dst_c)

                        for i, v in enumerate(values):
                            ws_tr.cell(row=start_row_for_columns + i, column=dst_c, value=v)

                        log.append(
                            f"  [m={m_idx}] {m_folder.name}: root {col_letter(src_c)}(:) -> "
                            f"sheet#{transient_sheet_index}('{ws_tr.title}') "
                            f"{dstL}{start_row_for_columns}:{dstL}{start_row_for_columns + len(values) - 1}"
                        )

            except Exception as e:
                log.append(f"  [m={m_idx}] {m_folder.name}: ERROR (root->transient) {e}")
                log.append("    " + traceback.format_exc().replace("\n", "\n    "))

        log.append("")

    # =========================================================
    # ✅【修正箇所】最後に、まとめシートを作成して転記
    # =========================================================
    try:
        summary_sheet_index = len(wb.worksheets) + 1
        ws_sum = ensure_sheet_by_index_and_rename(wb, summary_sheet_index, desired_title="まとめ")

        # まとめのデータ開始行：3行目
        data_start_row = 3
        n_count = len(n_folders)

        # -------------------------
        # jw側ヘッダー
        # ③ 1行目: mフォルダ名を2列目から2列間隔でm回
        # ② 2行目: "I | V" を2列目から2列間隔でm回
        # -------------------------
        for X in range(1, mmax_global + 1):
            cI = 2 * X          # 2X列
            cV = 2 * X + 1      # 2X+1列
            mname = m_names_by_index.get(X, f"m{X:02d}")

            ws_sum.cell(row=1, column=cI, value=mname)
            ws_sum.cell(row=2, column=cI, value="I")
            ws_sum.cell(row=2, column=cV, value="V")

        # -------------------------
        # 過渡側ヘッダー
        # 開始列 = 2mmax + 2
        # ③ 1行目: mフォルダ名を(2mmax+2)列目から2列間隔でm回
        # ② 2行目: "I | V" を同様に
        # -------------------------
        trans_start = 2 * mmax_global + 2  # 2mmax+2
        for X in range(1, mmax_global + 1):
            cI = trans_start + 2 * (X - 1)
            cV = cI + 1
            mname = m_names_by_index.get(X, f"m{X:02d}")

            ws_sum.cell(row=1, column=cI, value=mname)
            ws_sum.cell(row=2, column=cI, value="I")
            ws_sum.cell(row=2, column=cV, value="V")

        # -------------------------
        # データ転記
        # jw: E列とJ列 -> まとめ 3行目から 2X,2X+1列
        # 過渡: B列とE列 -> まとめ 3行目から 2mmax+2X, 2mmax+2X+1列
        # -------------------------
        for n_idx in range(1, n_count + 1):
            dst_row = data_start_row + (n_idx - 1)

            # jwのn行（ヘッダありなら2行目がn=1）
            src_row_jw = (base_row_for_single - 1) + n_idx  # 2,3,4...

            # 過渡のn行（ヘッダありなら2行目がn=1）
            src_row_tr = (2 if header_row else 1) + (n_idx - 1)

            # jw側：m番目シート(= m_idx) から E,J を取得
            for X in range(1, mmax_global + 1):
                ws_jw = ensure_sheet_by_index_and_rename(wb, X, desired_title=m_names_by_index.get(X, str(X)))

                cI = 2 * X
                cV = 2 * X + 1

                vE = ws_jw.cell(row=src_row_jw, column=5).value   # E
                vJ = ws_jw.cell(row=src_row_jw, column=10).value  # J

                ws_sum.cell(row=dst_row, column=cI, value=vE)
                ws_sum.cell(row=dst_row, column=cV, value=vJ)

            # 過渡側：mmax_global+m番目シート から B,E を取得
            for X in range(1, mmax_global + 1):
                ws_tr = ensure_sheet_by_index_and_rename(
                    wb,
                    mmax_global + X,
                    desired_title=f"{m_names_by_index.get(X, f'm{X:02d}')}_過渡(V3+V6)",
                )

                cI = trans_start + 2 * (X - 1)
                cV = cI + 1

                vB = ws_tr.cell(row=src_row_tr, column=2).value  # B
                vE = ws_tr.cell(row=src_row_tr, column=5).value  # E

                ws_sum.cell(row=dst_row, column=cI, value=vB)
                ws_sum.cell(row=dst_row, column=cV, value=vE)

        log.append("Created summary sheet: まとめ")
        log.append(f"  jw     -> start col 2, pairs (I,V) repeated mmax_global={mmax_global}")
        log.append(f"  transient -> start col {trans_start}, pairs (I,V) repeated mmax_global={mmax_global}")
        log.append("")

    except Exception as e:
        log.append(f"ERROR (summary sheet) {e}")
        log.append("    " + traceback.format_exc().replace("\n", "\n    "))

    # 既定Sheetが空で残る場合は消す（必要なら）
    # if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
    #     del wb["Sheet"]

    wb.save(output_xlsx)
    log_path.write_text("\n".join(log), encoding="utf-8")
    return output_xlsx


def gui_main():
    if tk is None:
        raise RuntimeError("tkinter is not available.")

    root = tk.Tk()
    root.withdraw()

    messagebox.showinfo("No-load aggregator", "A（親フォルダ）を選択してください。")
    A = filedialog.askdirectory(title="Select A folder")
    if not A:
        return

    header = messagebox.askyesno(
        "No-load aggregator",
        "Excelの1行目はヘッダですか？（はい→2行目から書き込み）"
    )

    out = filedialog.asksaveasfilename(
        title="Save output xlsx as",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
    )

    out_path = run(
        Path(A),
        output_xlsx=Path(out) if out else None,
        header_row=header,
    )

    messagebox.showinfo(
        "No-load aggregator",
        f"完了\n出力: {out_path}\nログ: {out_path.with_suffix('.log.txt')}"
    )


if __name__ == "__main__":
    # CLI:
    # python main.py "C:\path\to\A" "C:\path\to\out.xlsx" 1
    if len(sys.argv) >= 2:
        A = Path(sys.argv[1])
        out = Path(sys.argv[2]) if len(sys.argv) >= 3 else None
        header = bool(int(sys.argv[3])) if len(sys.argv) >= 4 else True

        out_path = run(
            A,
            output_xlsx=out,
            header_row=header,
        )
        print(f"Done: {out_path}")
        print(f"Log : {out_path.with_suffix('.log.txt')}")
    else:
        gui_main()
