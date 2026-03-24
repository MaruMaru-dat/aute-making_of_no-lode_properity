"""
Table manipulation helpers for aggregating CSV data into Excel workbooks.

This module wraps ``openpyxl`` to load workbooks, create or rename worksheets,
and provides utilities to read specific cells or entire columns from CSV files.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Optional, Tuple, List

import openpyxl
from openpyxl.workbook.workbook import Workbook


def col_letter(col_1based: int) -> str:
    """Convert a 1‑based column index to its Excel column letter.

    For example, ``1`` -> ``'A'``, ``27`` -> ``'AA'``.

    Parameters
    ----------
    col_1based : int
        1‑based column index.

    Returns
    -------
    str
        Corresponding Excel column letter.
    """
    s = ""
    n = col_1based
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def read_csv_cell_1based(csv_path: Path, row_1based: int, col_1based: int) -> Tuple[Optional[float], str]:
    """Read a single cell from a CSV as a float.

    Attempts to decode the CSV using several encodings. If the cell content is
    empty or cannot be converted to a float, returns ``None`` along with the
    original string.

    Parameters
    ----------
    csv_path : Path
        Path to the CSV file.
    row_1based : int
        1‑based row index.
    col_1based : int
        1‑based column index.

    Returns
    -------
    Tuple[Optional[float], str]
        Tuple of the parsed float value (or ``None``) and the raw string.
    """
    encodings = ["utf-8-sig", "cp932"]
    last_err = None

    for enc in encodings:
        try:
            text = csv_path.read_text(encoding=enc, errors="strict")
            try:
                dialect = csv.Sniffer().sniff(text[:4096])
            except Exception:
                dialect = csv.excel
                dialect.delimiter = ","

            rows = list(csv.reader(text.splitlines(), dialect))
            r_idx = row_1based - 1
            c_idx = col_1based - 1

            if r_idx < 0 or c_idx < 0 or r_idx >= len(rows):
                return None, ""

            row = rows[r_idx]
            if c_idx >= len(row):
                return None, ""

            raw = row[c_idx].strip()
            if raw == "" or raw.lower() in {"nan", "none", "null"}:
                return None, raw

            try:
                return float(raw), raw
            except Exception:
                # Remove thousands separators and retry
                raw2 = raw.replace(",", "")
                try:
                    return float(raw2), raw
                except Exception:
                    return None, raw

        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(f"Failed to read CSV: {csv_path}\n{last_err}")


def read_csv_column_1based(csv_path: Path, col_1based: int) -> List[Optional[float]]:
    """Read an entire column from a CSV file.

    Non‑numeric or empty values are returned as ``None``. The function attempts
    several encodings when reading the file. It does not attempt to skip
    header rows; callers should handle any necessary slicing.

    Parameters
    ----------
    csv_path : Path
        Path to the CSV file.
    col_1based : int
        1‑based column index to read.

    Returns
    -------
    List[Optional[float]]
        List of parsed float values, with ``None`` for cells that cannot be
        interpreted as floats.
    """
    encodings = ["utf-8-sig", "cp932"]
    last_err = None

    for enc in encodings:
        try:
            text = csv_path.read_text(encoding=enc, errors="strict")
            try:
                dialect = csv.Sniffer().sniff(text[:4096])
            except Exception:
                dialect = csv.excel
                dialect.delimiter = ","

            rows = list(csv.reader(text.splitlines(), dialect))
            c_idx = col_1based - 1

            out: List[Optional[float]] = []
            for r in rows:
                if c_idx >= len(r):
                    out.append(None)
                    continue
                raw = (r[c_idx] or "").strip()
                if raw == "" or raw.lower() in {"nan", "none", "null"}:
                    out.append(None)
                    continue
                try:
                    out.append(float(raw))
                except Exception:
                    raw2 = raw.replace(",", "")
                    try:
                        out.append(float(raw2))
                    except Exception:
                        out.append(None)

            return out

        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(f"Failed to read CSV column: {csv_path}\n{last_err}")


def ensure_sheet_by_index_and_rename(wb: Workbook, target_index_1based: int, desired_title: str):
    """Ensure the workbook has at least ``target_index_1based`` worksheets and rename it.

    If the sheet already exists but has a different title, it will be renamed.
    If the desired title conflicts with an existing sheet, a suffix is appended
    (e.g. ``_2``, ``_3``) to make it unique.

    Parameters
    ----------
    wb : Workbook
        The workbook to modify.
    target_index_1based : int
        1‑based index of the sheet to retrieve or create.
    desired_title : str
        Desired title for the sheet. If blank, the index is used as the title.

    Returns
    -------
    Worksheet
        The worksheet corresponding to ``target_index_1based``.
    """
    # Create sheets until the desired index exists
    while len(wb.worksheets) < target_index_1based:
        new_idx = len(wb.worksheets) + 1
        wb.create_sheet(title=str(new_idx))

    ws = wb.worksheets[target_index_1based - 1]

    title = (desired_title or "").strip()
    if title == "":
        title = str(target_index_1based)

    if ws.title != title:
        # If another sheet already has the desired title, append a numeric suffix
        if title in wb.sheetnames and wb[title] is not ws:
            i = 2
            base = title
            while f"{base}_{i}" in wb.sheetnames:
                i += 1
            title = f"{base}_{i}"
        ws.title = title

    return ws


def load_workbook(path: Path) -> Workbook:
    """Load an Excel workbook using openpyxl."""
    return openpyxl.load_workbook(path)