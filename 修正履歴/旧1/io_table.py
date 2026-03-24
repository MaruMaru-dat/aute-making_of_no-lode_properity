# io_table.py

from __future__ import annotations

import csv
from pathlib import Path
from typing import Optional, Tuple, List

import openpyxl
from openpyxl.workbook.workbook import Workbook

def col_letter(col_1based: int) -> str:
    s = ""
    n = col_1based
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def read_csv_cell_1based(csv_path: Path, row_1based: int, col_1based: int) -> Tuple[Optional[float], str]:
    encodings = ["utf-8-sig", "cp932"]
    last_err = None

    for enc in encodings:
        try:
            text = csv_path.read_text(encoding=enc, errors="strict")
            try:
                dialect = csv.Sniffer().sniff(text[:4096])
            except Exception:
                dialect = csv.excel
                dialect.delimiter = ","

            rows = list(csv.reader(text.splitlines(), dialect))
            r_idx = row_1based - 1
            c_idx = col_1based - 1

            if r_idx < 0 or c_idx < 0 or r_idx >= len(rows):
                return None, ""

            row = rows[r_idx]
            if c_idx >= len(row):
                return None, ""

            raw = row[c_idx].strip()
            if raw == "" or raw.lower() in {"nan", "none", "null"}:
                return None, raw

            try:
                return float(raw), raw
            except Exception:
                raw2 = raw.replace(",", "")
                try:
                    return float(raw2), raw
                except Exception:
                    return None, raw

        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(f"Failed to read CSV: {csv_path}\n{last_err}")

def read_csv_column_1based(csv_path: Path, col_1based: int) -> List[Optional[float]]:
    """
    CSVの指定「列」を上から全部読む（1始まり列番号）
    変換できないセルは None にする
    """
    encodings = ["utf-8-sig", "cp932"]
    last_err = None

    for enc in encodings:
        try:
            text = csv_path.read_text(encoding=enc, errors="strict")
            try:
                dialect = csv.Sniffer().sniff(text[:4096])
            except Exception:
                dialect = csv.excel
                dialect.delimiter = ","

            rows = list(csv.reader(text.splitlines(), dialect))
            c_idx = col_1based - 1

            out: List[Optional[float]] = []
            for r in rows:
                if c_idx >= len(r):
                    out.append(None)
                    continue
                raw = (r[c_idx] or "").strip()
                if raw == "" or raw.lower() in {"nan", "none", "null"}:
                    out.append(None)
                    continue
                try:
                    out.append(float(raw))
                except Exception:
                    raw2 = raw.replace(",", "")
                    try:
                        out.append(float(raw2))
                    except Exception:
                        out.append(None)

            return out

        except Exception as e:
            last_err = e
            continue

    raise RuntimeError(f"Failed to read CSV column: {csv_path}\n{last_err}")

def ensure_sheet_by_index_and_rename(wb: Workbook, target_index_1based: int, desired_title: str):
    while len(wb.worksheets) < target_index_1based:
        new_idx = len(wb.worksheets) + 1
        wb.create_sheet(title=str(new_idx))

    ws = wb.worksheets[target_index_1based - 1]

    title = (desired_title or "").strip()
    if title == "":
        title = str(target_index_1based)

    if ws.title != title:
        if title in wb.sheetnames and wb[title] is not ws:
            i = 2
            while f"{title}_{i}" in wb.sheetnames:
                i += 1
            title = f"{title}_{i}"
        ws.title = title

    return ws

def load_workbook(path: Path):
    return openpyxl.load_workbook(path)