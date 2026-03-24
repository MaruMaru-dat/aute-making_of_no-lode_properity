# main.py
from __future__ import annotations

import sys
import traceback
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Any

import openpyxl  # ★追加（テンプレを使わず新規作成するため）

from io_fs import list_subfolders_sorted, circuit_csv_path, circuit_csv_path_root
from io_table import (
    ensure_sheet_by_index_and_rename,
    read_csv_cell_1based,
    read_csv_column_1based,
    col_letter,
)

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception:
    tk = None


# -------------------------
# ★自動ヘッダー定義
# -------------------------

JW_HEADERS = [
    "U相電源電圧(6)",
    "Vabs",
    "Vの線間電圧(実効値)",
    "Iabs",
    "Iの相電流(実効値)",
    "", "", "",
    "Vabs(3+6)",
    "Vの線間電圧(実効値)",
]

TRANSIENT_HEADERS = [
    "U相電源電圧(6)",
    "Iの実効値",
    "",
    "Vの実効値",
    "Vの線間電圧",
]


def _write_row(ws, row: int, values: List[str], start_col: int = 1):
    """1行分を左から順に書く（既存でも上書き）"""
    for j, v in enumerate(values):
        ws.cell(row=row, column=start_col + j, value=v)


def run(
    A: Path,
    output_xlsx: Optional[Path] = None,
    header_row: bool = True,
    read_to_write_map: Optional[List[Dict[str, Any]]] = None,
) -> Path:
    """
    ✅テンプレ不要版：
    - Excelは新規作成（openpyxl.Workbook）
    - jwシート/過渡シートのヘッダーを自動生成
    """
    A = A.resolve()

    if read_to_write_map is None:
        read_to_write_map = [
            {"src_row": 8, "src_col": 4, "dst_col": 4, "label": "D8"},  # jw: D8 -> D
            {"src_row": 8, "src_col": 8, "dst_col": 2, "label": "H8"},  # jw: H8 -> B
        ]

    if output_xlsx is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_xlsx = Path.cwd() / f"無負荷特性_全自動化_out_{stamp}.xlsx"
    else:
        output_xlsx = output_xlsx.resolve()

    if not A.is_dir():
        raise FileNotFoundError(f"A folder not found: {A}")

    # ★テンプレではなく新規ブック
    wb = openpyxl.Workbook()

    # Logging
    log_path = output_xlsx.with_suffix(".log.txt")
    log: List[str] = []
    log.append(f"Run at: {datetime.now().isoformat(timespec='seconds')}")
    log.append(f"A: {A}")
    log.append(f"Output: {output_xlsx}")
    log.append("")

    n_folders = list_subfolders_sorted(A)
    log.append(f"Found n folders: {len(n_folders)}")
    log.append("")

    # ★ヘッダーありなら n=1 は2行目から
    base_row_for_single = 2 if header_row else 1

    # 列貼り（過渡列ベタ貼り）の開始行（あなたの運用通り）
    start_row_for_columns = 2 if header_row else 1

    # 過渡シートのヘッダー（1行目/2行目）開始列：M列
    transient_header_start_col = 13  # M = 13

    # RMS式の参照範囲（固定指定）
    rms_r1 = 1283
    rms_r2 = 1538
    rms_N = 256

    for n_idx, n_folder in enumerate(n_folders, start=1):
        m_folders = list_subfolders_sorted(n_folder)
        log.append(f"[n={n_idx}/{len(n_folders)}] {n_folder.name} -> m folders: {len(m_folders)}")

        target_row_single = (base_row_for_single - 1) + n_idx  # headerあり: 2,3,4...

        for m_idx, m_folder in enumerate(m_folders, start=1):
            # -------------------------
            # sheet#1..mmax は「定常」(ABH / Original / ...)
            # -------------------------
            sheet_index = m_idx
            ws = ensure_sheet_by_index_and_rename(wb, sheet_index, desired_title=m_folder.name)

            # ★(①) jwシートのヘッダーを自動生成
            if header_row:
                _write_row(ws, row=1, values=JW_HEADERS, start_col=1)

            # ★(あなたの要望) jwシートのA列に nフォルダ名
            ws.cell(row=target_row_single, column=1, value=n_folder.name)

            # -------------------------
            # ★ jwシート：数式を自動入力（修正箇所）
            #   C列: =B{r}/SQRT(2)*SQRT(3)
            #   E列: =D{r}/SQRT(2)
            #   I列: =B{r}+G{r}
            #   J列: =I{r}/SQRT(2)*SQRT(3)
            # -------------------------
            r = target_row_single
            ws.cell(row=r, column=3, value=f"=B{r}/SQRT(2)*SQRT(3)")  # C
            ws.cell(row=r, column=5, value=f"=D{r}/SQRT(2)")         # E
            ws.cell(row=r, column=9, value=f"=B{r}+G{r}")            # I
            ws.cell(row=r, column=10, value=f"=I{r}/SQRT(2)*SQRT(3)") # J
            # -------------------------
            # (A) jw/circuit.csv：単点貼り
            # -------------------------
            jw_csv = circuit_csv_path(n_folder, m_folder, jw_folder_name="jw", csv_name="circuit.csv")
            try:
                if not jw_csv.exists():
                    log.append(f"  [m={m_idx}] {m_folder.name}: MISSING {jw_csv}")
                else:
                    for item in read_to_write_map:
                        src_r = int(item["src_row"])
                        src_c = int(item["src_col"])
                        dst_c = int(item["dst_col"])
                        label = str(item.get("label", f"R{src_r}C{src_c}"))

                        val, raw = read_csv_cell_1based(jw_csv, row_1based=src_r, col_1based=src_c)
                        dstL = col_letter(dst_c)

                        ws.cell(row=target_row_single, column=dst_c, value=val if val is not None else None)

                        if val is None:
                            log.append(
                                f"  [m={m_idx}] {m_folder.name}: {label}(raw='{raw}') -> "
                                f"sheet#{sheet_index}('{ws.title}') {dstL}{target_row_single}=BLANK"
                            )
                        else:
                            log.append(
                                f"  [m={m_idx}] {m_folder.name}: {label}={val} -> "
                                f"sheet#{sheet_index}('{ws.title}') {dstL}{target_row_single}"
                            )

            except Exception as e:
                log.append(f"  [m={m_idx}] {m_folder.name}: ERROR (jw) {e}")
                log.append("    " + traceback.format_exc().replace("\n", "\n    "))

            # -------------------------
            # (B) m直下 circuit.csv：列貼り + 過渡シート
            # -------------------------
            root_csv = circuit_csv_path_root(m_folder, csv_name="circuit.csv")
            try:
                if not root_csv.exists():
                    log.append(f"  [m={m_idx}] {m_folder.name}: MISSING {root_csv}")
                else:
                    mmax = len(m_folders)

                    transient_sheet_index = mmax + m_idx
                    transient_title = f"{m_folder.name}_過渡(V3+V6)"
                    ws_tr = ensure_sheet_by_index_and_rename(
                        wb, transient_sheet_index, desired_title=transient_title
                    )

                    # ★(②) 過渡シートのヘッダーを自動生成（A1〜E1）
                    if header_row:
                        _write_row(ws_tr, row=1, values=TRANSIENT_HEADERS, start_col=1)

                    base_dst = transient_header_start_col + 3 * (n_idx - 1)  # M + 3*(n-1)
                    dst_cols = [base_dst, base_dst + 1, base_dst + 2]

                    src_cols = [8, 17, 20]  # H, Q, T（1始まり）

                    summary_row = (2 if header_row else 1) + (n_idx - 1)
                   
                    # -------------------------
                    # ★ 過渡シート：数式を自動入力（修正箇所）
                    #   E列: =D{r}*SQRT(3)
                    # -------------------------
                    r2 = summary_row
                    ws_tr.cell(row=r2, column=5, value=f"=D{r2}*SQRT(3)")  # E

                    # A列に nフォルダ名（過渡のみ）
                    ws_tr.cell(row=summary_row, column=1, value=n_folder.name)

                    # 1行目のM列から nフォルダ名を3列間隔（既存仕様）
                    ws_tr.cell(row=1, column=base_dst, value=n_folder.name)

                    # 2行目のM列から「I」「V3」「V6」
                    ws_tr.cell(row=2, column=base_dst, value="I")
                    ws_tr.cell(row=2, column=base_dst + 1, value="V3")
                    ws_tr.cell(row=2, column=base_dst + 2, value="V6")

                    # B列 RMS（X=M/P/S...）
                    X = col_letter(base_dst)
                    ws_tr.cell(
                        row=summary_row,
                        column=2,
                        value=f"=SQRT(SUMSQ(${X}${rms_r1}:${X}${rms_r2})/{rms_N})",
                    )

                    # D列 RMS（Y=N/Q/T..., Z=O/R/U...）
                    Y = col_letter(base_dst + 1)
                    Z = col_letter(base_dst + 2)
                    ws_tr.cell(
                        row=summary_row,
                        column=4,
                        value=(
                            f"=SQRT(SUMSQ(${Y}${rms_r1}:${Y}${rms_r2})/{rms_N})"
                            f"+SQRT(SUMSQ(${Z}${rms_r1}:${Z}${rms_r2})/{rms_N})"
                        ),
                    )

                    # 「列を丸ごと貼り」
                    for src_c, dst_c in zip(src_cols, dst_cols):
                        values = read_csv_column_1based(root_csv, col_1based=src_c)
                        dstL = col_letter(dst_c)

                        for i, v in enumerate(values):
                            ws_tr.cell(row=start_row_for_columns + i, column=dst_c, value=v)

                        log.append(
                            f"  [m={m_idx}] {m_folder.name}: root {col_letter(src_c)}(:) -> "
                            f"sheet#{transient_sheet_index}('{ws_tr.title}') "
                            f"{dstL}{start_row_for_columns}:{dstL}{start_row_for_columns + len(values) - 1}"
                        )

            except Exception as e:
                log.append(f"  [m={m_idx}] {m_folder.name}: ERROR (root->transient) {e}")
                log.append("    " + traceback.format_exc().replace("\n", "\n    "))

        log.append("")

    # ★openpyxl新規ブックの「不要な既定シート」を消したい場合（任意）
    # 既にsheet#1へリネームされていることが多いので、基本不要。
    # もし空のシートが残るなら、以下を有効化して調整してください。
    # if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
    #     del wb["Sheet"]

    wb.save(output_xlsx)
    log_path.write_text("\n".join(log), encoding="utf-8")
    return output_xlsx


def gui_main():
    if tk is None:
        raise RuntimeError("tkinter is not available.")

    root = tk.Tk()
    root.withdraw()

    messagebox.showinfo("No-load aggregator", "A（親フォルダ）を選択してください。")
    A = filedialog.askdirectory(title="Select A folder")
    if not A:
        return

    header = messagebox.askyesno(
        "No-load aggregator",
        "Excelの1行目はヘッダですか？（はい→2行目から書き込み）"
    )

    # ★保存先を聞く（キャンセルなら自動命名）
    out = filedialog.asksaveasfilename(
        title="Save output xlsx as",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
    )
    out_path = run(
        Path(A),
        output_xlsx=Path(out) if out else None,
        header_row=header,
    )

    messagebox.showinfo(
        "No-load aggregator",
        f"完了\n出力: {out_path}\nログ: {out_path.with_suffix('.log.txt')}"
    )


if __name__ == "__main__":
    # CLI:
    # python main.py "C:\path\to\A" "C:\path\to\out.xlsx" 1
    if len(sys.argv) >= 2:
        A = Path(sys.argv[1])
        out = Path(sys.argv[2]) if len(sys.argv) >= 3 else None
        header = bool(int(sys.argv[3])) if len(sys.argv) >= 4 else True

        out_path = run(
            A,
            output_xlsx=out,
            header_row=header,
        )
        print(f"Done: {out_path}")
        print(f"Log : {out_path.with_suffix('.log.txt')}")
    else:
        gui_main()
