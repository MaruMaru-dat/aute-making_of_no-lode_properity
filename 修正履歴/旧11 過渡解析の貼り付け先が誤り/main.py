from __future__ import annotations

import sys
import traceback
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Dict, Any

import math
import openpyxl  # テンプレ不要の新規作成

from io_fs import list_subfolders_sorted, circuit_csv_path, circuit_csv_path_root
from io_table import (
    ensure_sheet_by_index_and_rename,
    read_csv_cell_1based,
    read_csv_column_1based,
    col_letter,
)

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except Exception:
    tk = None


# -------------------------
# 自動ヘッダー定義
# -------------------------
JW_HEADERS = [
    "U相電源電圧(6)",            # A
    "Vabs",                      # B
    "Vの線間電圧(実効値)",         # C (=B/√2*√3)
    "Iabs",                      # D
    "Iの相電流(実効値)",           # E (=D/√2)
    "",                          # F
    "",                          # G ← ★ここに jw/circuit.csv の H3 を入れる（今回の修正）
    "",                          # H
    "Vabs(3+6)",                 # I (=B+G)
    "Vの線間電圧(実効値)",         # J (=I/√2*√3)
]

TRANSIENT_HEADERS = [
    "U相電源電圧(6)",  # A
    "Iの実効値",       # B（RMSの式）
    "",                # C
    "Vの実効値",       # D（RMSの式）
    "Vの線間電圧",     # E (=D*√3)
]


def _write_row(ws, row: int, values: List[str], start_col: int = 1):
    """1行分を左から順に書く（既存でも上書き）"""
    for j, v in enumerate(values):
        ws.cell(row=row, column=start_col + j, value=v)


def _to_float_or_none(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    try:
        return float(str(x).strip())
    except Exception:
        return None


def _rms_from_sheet_column(ws, col_1based: int, r1: int, r2: int):
    """ws の指定列・指定行範囲から RMS を数値で計算"""
    s = 0.0
    cnt = 0
    for r in range(r1, r2 + 1):
        v = _to_float_or_none(ws.cell(row=r, column=col_1based).value)
        if v is None:
            continue
        s += v * v
        cnt += 1
    if cnt == 0:
        return None
    return math.sqrt(s / cnt)


def run(
    A: Path,
    output_xlsx: Optional[Path] = None,
    header_row: bool = True,
    read_to_write_map: Optional[List[Dict[str, Any]]] = None,
) -> Path:
    """
    ✅テンプレ不要版：
    - Excelは新規作成（openpyxl.Workbook）
    - jwシート/過渡シートのヘッダーを自動生成
    - jw/過渡シートには「数式」を入れる（従来通り）
    - 最後に「まとめ」シートを作り、jw/過渡の値を「数値」で貼る
    - ★root_csv が無い場合は「過渡シートを作らない」方針（今回の指定）
    - ★jwシートのG列に jw/circuit.csv の H3 を格納（今回の修正）
    """
    A = A.resolve()

    # ★ jw/circuit.csv から貼る値
    # - D8 -> D列（Iabs）
    # - H8 -> B列（Vabs）
    # - H3 -> G列（Vabs(3+6) の加算用：今回の修正）
    if read_to_write_map is None:
        read_to_write_map = [
            {"src_row": 8, "src_col": 4, "dst_col": 4, "label": "D8"},   # -> D
            {"src_row": 8, "src_col": 8, "dst_col": 2, "label": "H8"},   # -> B
            {"src_row": 3, "src_col": 8, "dst_col": 7, "label": "H3"},   # -> G（★追加）
        ]

    if output_xlsx is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_xlsx = Path.cwd() / f"無負荷特性_全自動化_out_{stamp}.xlsx"
    else:
        output_xlsx = output_xlsx.resolve()

    if not A.is_dir():
        raise FileNotFoundError(f"A folder not found: {A}")

    wb = openpyxl.Workbook()

    # Logging
    log_path = output_xlsx.with_suffix(".log.txt")
    log: List[str] = []
    log.append(f"Run at: {datetime.now().isoformat(timespec='seconds')}")
    log.append(f"A: {A}")
    log.append(f"Output: {output_xlsx}")
    log.append("")

    n_folders = list_subfolders_sorted(A)
    log.append(f"Found n folders: {len(n_folders)}")
    log.append("")

    # ヘッダーありなら n=1 は2行目から
    base_row_for_single = 2 if header_row else 1

    # 列貼り（過渡列ベタ貼り）の開始行
    start_row_for_columns = 2 if header_row else 1

    # 過渡シートのヘッダー開始列：M列
    transient_header_start_col = 13  # M = 13

    # RMS の参照範囲（固定指定）
    rms_r1 = 1283
    rms_r2 = 1538
    rms_N = 256  # シート内の数式用

    sqrt2 = math.sqrt(2.0)
    sqrt3 = math.sqrt(3.0)

    # =========================================================
    # 事前スキャン：mmax_global と mフォルダ名を確定（まとめ用）
    # =========================================================
    mmax_global = 0
    m_names_by_index: Dict[int, str] = {}  # m_idx(1-based)->name

    for n_folder in n_folders:
        m_folders_tmp = list_subfolders_sorted(n_folder)
        mmax_global = max(mmax_global, len(m_folders_tmp))
        for i, mf in enumerate(m_folders_tmp, start=1):
            if i not in m_names_by_index:
                m_names_by_index[i] = mf.name

    log.append(f"Detected mmax_global: {mmax_global}")
    log.append("")

    # =========================================================
    # 本処理：jw/過渡シートを作る
    # =========================================================
    for n_idx, n_folder in enumerate(n_folders, start=1):
        m_folders = list_subfolders_sorted(n_folder)
        log.append(f"[n={n_idx}/{len(n_folders)}] {n_folder.name} -> m folders: {len(m_folders)}")

        target_row_single = (base_row_for_single - 1) + n_idx  # 2,3,4...

        for m_idx, m_folder in enumerate(m_folders, start=1):
            # -------------------------
            # jwシート（定常）
            # -------------------------
            sheet_index = m_idx
            ws = ensure_sheet_by_index_and_rename(wb, sheet_index, desired_title=m_folder.name)

            if header_row:
                _write_row(ws, row=1, values=JW_HEADERS, start_col=1)

            # A列：nフォルダ名
            ws.cell(row=target_row_single, column=1, value=n_folder.name)

            # jwシート：数式（従来通り）
            r = target_row_single
            ws.cell(row=r, column=3, value=f"=B{r}/SQRT(2)*SQRT(3)")     # C
            ws.cell(row=r, column=5, value=f"=D{r}/SQRT(2)")            # E
            ws.cell(row=r, column=9, value=f"=B{r}+G{r}")               # I（★Gを使う）
            ws.cell(row=r, column=10, value=f"=I{r}/SQRT(2)*SQRT(3)")   # J

            # jw/circuit.csv：単点貼り
            jw_csv = circuit_csv_path(n_folder, m_folder, jw_folder_name="jw", csv_name="circuit.csv")
            try:
                if not jw_csv.exists():
                    log.append(f"  [m={m_idx}] {m_folder.name}: MISSING {jw_csv}")
                else:
                    for item in read_to_write_map:
                        src_r = int(item["src_row"])
                        src_c = int(item["src_col"])
                        dst_c = int(item["dst_col"])
                        label = str(item.get("label", f"R{src_r}C{src_c}"))

                        val, raw = read_csv_cell_1based(jw_csv, row_1based=src_r, col_1based=src_c)
                        dstL = col_letter(dst_c)

                        ws.cell(row=target_row_single, column=dst_c, value=val if val is not None else None)

                        if val is None:
                            log.append(
                                f"  [m={m_idx}] {m_folder.name}: {label}(raw='{raw}') -> "
                                f"sheet#{sheet_index}('{ws.title}') {dstL}{target_row_single}=BLANK"
                            )
                        else:
                            log.append(
                                f"  [m={m_idx}] {m_folder.name}: {label}={val} -> "
                                f"sheet#{sheet_index}('{ws.title}') {dstL}{target_row_single}"
                            )
            except Exception as e:
                log.append(f"  [m={m_idx}] {m_folder.name}: ERROR (jw) {e}")
                log.append("    " + traceback.format_exc().replace("\n", "\n    "))

            # -------------------------
            # 過渡シート（★root_csv が無いなら作らない方針）
            # -------------------------
            root_csv = circuit_csv_path_root(m_folder, csv_name="circuit.csv")
            try:
                if not root_csv.exists():
                    # ★今回の方針：無いなら過渡シートは作成しない
                    log.append(f"  [m={m_idx}] {m_folder.name}: root missing -> transient sheet NOT created")
                else:
                    transient_title = f"{m_folder.name}_過渡(V3+V6)"
                    if transient_title in wb.sheetnames:
                        ws_tr = wb[transient_title]
                    else:
                        ws_tr = wb.create_sheet(title=transient_title)

                    if header_row:
                        _write_row(ws_tr, row=1, values=TRANSIENT_HEADERS, start_col=1)

                    base_dst = transient_header_start_col + 3 * (n_idx - 1)  # M + 3*(n-1)
                    dst_cols = [base_dst, base_dst + 1, base_dst + 2]
                    src_cols = [8, 17, 20]  # H, Q, T

                    summary_row = (2 if header_row else 1) + (n_idx - 1)

                    # E列 = D*√3（従来通り）
                    r2 = summary_row
                    ws_tr.cell(row=r2, column=5, value=f"=D{r2}*SQRT(3)")

                    # A列：nフォルダ名（過渡）
                    ws_tr.cell(row=summary_row, column=1, value=n_folder.name)

                    # 1行目：n名（波形列の先頭に表示）
                    ws_tr.cell(row=1, column=base_dst, value=n_folder.name)

                    # 2行目：I, V3, V6
                    ws_tr.cell(row=2, column=base_dst, value="I")
                    ws_tr.cell(row=2, column=base_dst + 1, value="V3")
                    ws_tr.cell(row=2, column=base_dst + 2, value="V6")

                    # B列 RMS（X=M/P/S...）
                    Xcol = col_letter(base_dst)
                    ws_tr.cell(
                        row=summary_row,
                        column=2,
                        value=f"=SQRT(SUMSQ(${Xcol}${rms_r1}:${Xcol}${rms_r2})/{rms_N})",
                    )

                    # D列 RMS（Y=N/Q/T..., Z=O/R/U...）
                    Ycol = col_letter(base_dst + 1)
                    Zcol = col_letter(base_dst + 2)
                    ws_tr.cell(
                        row=summary_row,
                        column=4,
                        value=(
                            f"=SQRT(SUMSQ(${Ycol}${rms_r1}:${Ycol}${rms_r2})/{rms_N})"
                            f"+SQRT(SUMSQ(${Zcol}${rms_r1}:${Zcol}${rms_r2})/{rms_N})"
                        ),
                    )

                    # 列を丸ごと貼り（波形）
                    for src_c, dst_c in zip(src_cols, dst_cols):
                        values = read_csv_column_1based(root_csv, col_1based=src_c)
                        dstL = col_letter(dst_c)

                        for i, v in enumerate(values):
                            ws_tr.cell(row=start_row_for_columns + i, column=dst_c, value=v)

                        log.append(
                            f"  [m={m_idx}] {m_folder.name}: root {col_letter(src_c)}(:) -> "
                            f"sheet('{ws_tr.title}') "
                            f"{dstL}{start_row_for_columns}:{dstL}{start_row_for_columns + len(values) - 1}"
                        )
            except Exception as e:
                log.append(f"  [m={m_idx}] {m_folder.name}: ERROR (root->transient) {e}")
                log.append("    " + traceback.format_exc().replace("\n", "\n    "))

        log.append("")

    # =========================================================
    # 最後に：まとめシートを作成（数値貼り）
    # =========================================================
    try:
        summary_title = "まとめ"
        if summary_title in wb.sheetnames:
            ws_sum = wb[summary_title]
        else:
            ws_sum = wb.create_sheet(title=summary_title)

        data_start_row = 3
        n_count = len(n_folders)

        # 1列目：n名（任意）
        ws_sum.cell(row=1, column=1, value="")
        ws_sum.cell(row=2, column=1, value="")

        # -------------------------
        # jw側ヘッダ
        # 1行目：m名，2行目：I,V（2列目から2列間隔）
        # -------------------------
        for X in range(1, mmax_global + 1):
            cI = 2 * X
            cV = 2 * X + 1
            mname = m_names_by_index.get(X, f"m{X:02d}")
            ws_sum.cell(row=1, column=cI, value=mname)
            ws_sum.cell(row=2, column=cI, value="I")
            ws_sum.cell(row=2, column=cV, value="V")

        # -------------------------
        # 過渡側ヘッダ
        # 1行目：m名，2行目：I,V（開始列=2mmax+2 から2列間隔）
        # -------------------------
        trans_start = 2 * mmax_global + 2  # 2mmax+2
        for X in range(1, mmax_global + 1):
            cI = trans_start + 2 * (X - 1)
            cV = cI + 1
            mname = m_names_by_index.get(X, f"m{X:02d}")
            ws_sum.cell(row=1, column=cI, value=mname)
            ws_sum.cell(row=2, column=cI, value="I")
            ws_sum.cell(row=2, column=cV, value="V")

        # -------------------------
        # データ（数値）貼り付け
        # jw: E列とJ列相当を「数値計算」して貼る（openpyxlは式を評価しないため）
        # 過渡: B列とE列相当を「RMS計算＋√3」で数値貼り
        # ※過渡シートが無いmは空欄のまま（あなたの方針）
        # -------------------------
        for n_idx in range(1, n_count + 1):
            dst_row = data_start_row + (n_idx - 1)

            # n名（任意）
            ws_sum.cell(row=dst_row, column=1, value=n_folders[n_idx - 1].name)

            # jwの行（ヘッダありなら2行目がn=1）
            src_row_jw = (base_row_for_single - 1) + n_idx

            # 過渡の波形列（n_idxに依存：M, N, O / P, Q, R ...）
            base_dst = transient_header_start_col + 3 * (n_idx - 1)
            col_I = base_dst
            col_V3 = base_dst + 1
            col_V6 = base_dst + 2

            # jw側（m=1..mmax_global）
            for X in range(1, mmax_global + 1):
                mname = m_names_by_index.get(X, f"m{X:02d}")
                if mname not in wb.sheetnames:
                    continue
                ws_jw = wb[mname]

                cI = 2 * X
                cV = 2 * X + 1

                # I（=E相当） = D/√2
                D_val = _to_float_or_none(ws_jw.cell(row=src_row_jw, column=4).value)  # D
                I_num = (D_val / sqrt2) if D_val is not None else None

                # V（=J相当） = (B+G)/√2*√3
                B_val = _to_float_or_none(ws_jw.cell(row=src_row_jw, column=2).value)  # B
                G_val = _to_float_or_none(ws_jw.cell(row=src_row_jw, column=7).value)  # G（★H3が入る）
                V_num = ((B_val + G_val) / sqrt2 * sqrt3) if (B_val is not None and G_val is not None) else None

                ws_sum.cell(row=dst_row, column=cI, value=I_num)
                ws_sum.cell(row=dst_row, column=cV, value=V_num)

            # 過渡側（m=1..mmax_global）
            for X in range(1, mmax_global + 1):
                mname = m_names_by_index.get(X, f"m{X:02d}")
                tr_name = f"{mname}_過渡(V3+V6)"
                if tr_name not in wb.sheetnames:
                    # ★無い場合は作らない方針：まとめも空欄のまま
                    continue
                ws_tr = wb[tr_name]

                cI = trans_start + 2 * (X - 1)
                cV = cI + 1

                # I = RMS(I列)
                I_rms = _rms_from_sheet_column(ws_tr, col_I, rms_r1, rms_r2)

                # V = (RMS(V3)+RMS(V6))*√3
                V3_rms = _rms_from_sheet_column(ws_tr, col_V3, rms_r1, rms_r2)
                V6_rms = _rms_from_sheet_column(ws_tr, col_V6, rms_r1, rms_r2)
                V_ll = ((V3_rms + V6_rms) * sqrt3) if (V3_rms is not None and V6_rms is not None) else None

                ws_sum.cell(row=dst_row, column=cI, value=I_rms)
                ws_sum.cell(row=dst_row, column=cV, value=V_ll)

        log.append("Created summary sheet: まとめ (values only, transient sheets optional)")
        log.append("")

    except Exception as e:
        log.append(f"ERROR (summary sheet) {e}")
        log.append("    " + traceback.format_exc().replace("\n", "\n    "))

    # 既定Sheetが残るなら削除（任意）
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        try:
            del wb["Sheet"]
        except Exception:
            pass

    wb.save(output_xlsx)
    log_path.write_text("\n".join(log), encoding="utf-8")
    return output_xlsx


def gui_main():
    if tk is None:
        raise RuntimeError("tkinter is not available.")

    root = tk.Tk()
    root.withdraw()

    messagebox.showinfo("No-load aggregator", "A（親フォルダ）を選択してください。")
    A = filedialog.askdirectory(title="Select A folder")
    if not A:
        return

    header = messagebox.askyesno(
        "No-load aggregator",
        "Excelの1行目はヘッダですか？（はい→2行目から書き込み）"
    )

    out = filedialog.asksaveasfilename(
        title="Save output xlsx as",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
    )

    out_path = run(
        Path(A),
        output_xlsx=Path(out) if out else None,
        header_row=header,
    )

    messagebox.showinfo(
        "No-load aggregator",
        f"完了\n出力: {out_path}\nログ: {out_path.with_suffix('.log.txt')}"
    )


if __name__ == "__main__":
    # CLI:
    # python main.py "C:\path\to\A" "C:\path\to\out.xlsx" 1
    if len(sys.argv) >= 2:
        A = Path(sys.argv[1])
        out = Path(sys.argv[2]) if len(sys.argv) >= 3 else None
        header = bool(int(sys.argv[3])) if len(sys.argv) >= 4 else True

        out_path = run(
            A,
            output_xlsx=out,
            header_row=header,
        )
        print(f"Done: {out_path}")
        print(f"Log : {out_path.with_suffix('.log.txt')}")
    else:
        gui_main()
